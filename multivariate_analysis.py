"""

Author: Lory H

Purpose: Univariate and multivariate ORs of Group 1 vs Group 2

"""


import pandas as pd
import numpy as np
#stats related
import scipy.stats as stats
import statsmodels.formula.api as smf
import statsmodels.api as sm
import scikit_posthocs as sp
import seaborn as sns
import matplotlib.pyplot as plt
#format related
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

"""Importing Excel sheets"""

path = "path name"

#Import data to dataframe
data = pd.read_excel (path+r'database.xlsx', sheet_name='Database') 
list(data.columns)

"""Functions"""

# Missing data summary
def missing_data(df, variable_list):
    records = []

    for var in variable_list:
        count = df[var].isna().sum()
        percent = (count / len(df)) * 100
        records.append({
            'variables': var,
            'missing_count': count,
            'missing_percent': percent
        })

    return pd.DataFrame(records)

# Getting variable types
def get_variable_types(df, threshold=2, cat_max_unique=10):
    binary = [col for col in df.columns if df[col].dropna().nunique() == 2]
    categorical = [col for col in df.columns if threshold < df[col].dropna().nunique() <= cat_max_unique and (pd.api.types.is_object_dtype(df[col]) or pd.api.types.is_numeric_dtype(df[col]))]
    continuous = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col]) and df[col].dropna().nunique() > cat_max_unique]
    return binary, categorical, continuous

# Converting list to dataframe
def ls_to_df(ls):
    new_df = []
    for df in ls :
        new_df += [pd.DataFrame(df)]
    return new_df

# Impute missing data
def impute_data(df):
    df_imp = df.copy()
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            mean_col = df[col].mean()
            df_imp[col] = df[col].fillna(mean_col)
        # Optionally skip non-numeric columns or use a different imputation strategy
    return df_imp, df_imp.isnull().sum()

# Identify significant variables
def significant_values (table, var, pval):
    sig_val = []
    for x in range(len(table[var])): 
        if table[pval].iloc[x]<0.05:
            sig_val += [table[var][x]]
    return sig_val

#Logistic regression for odds ratios and CI   
def log_model(data, endog, exog_list, name):
    try:
        X = data[exog_list]
        y = data[endog]

        X = sm.add_constant(X)  # add intercept
        model = sm.Logit(y, X).fit(method='newton', disp=0)

        # Get odds ratios and confidence intervals
        params = model.params
        conf = model.conf_int()
        pvals = model.pvalues

        # Exclude 'const' from the output
        params = params.drop('const', errors='ignore')
        conf = conf.drop(index='const', errors='ignore')
        pvals = pvals.drop(index='const', errors='ignore')

        or_df = pd.DataFrame(np.exp(params), columns=[f'OR_{name}'])
        or_df[f'p-value_{name}'] = pvals
        or_df[[f'2.5%_{name}', f'97.5%_{name}']] = np.exp(conf)

        return or_df

    except Exception as e:
        print(f"Model failed for variables {exog_list}: {e}")
        return pd.DataFrame()

""" Define type of variables to include in the analysis"""

# Use function to create your variable lists
data = data.drop(columns='patient_id')

binary_vars, categorical_vars, continuous_vars = get_variable_types(data)

# Delete categorical variables that include string (text)
categorical_vars = [col for col in categorical_vars if not pd.api.types.is_object_dtype(data[col])]

print(binary_vars,categorical_vars,continuous_vars)

# Define your dependant (outcome) and independant (population) variables

dep_bin = ['group name']

indep_bin = [col for col in binary_vars if col not in dep_bin]
indep_cat = categorical_vars
indep_cont = continuous_vars

"""Calculate missing data and percentages"""

# Missing data

    # calculate missing data and generate dataframes
results = []
for var in [binary_vars, categorical_vars, continuous_vars]:
    results += [missing_data(data, var)]

missing_binary_df, missing_categorical_df, missing_continuous_df = results

    # combine the results into one dataframe
missing_df = pd.concat([missing_binary_df, missing_categorical_df, missing_continuous_df])

"""Univariate and Multivariate analysis : odds ratios"""

# Create df with imputed missing data
data_imp, missing_counts = impute_data(data)


# Identify significant descriptive variables
desc_cat = categorical test results #from comparison.py outcomes
desc_cont = continuous test results #from comparison.py outcomes

    #print significant variables into a list
sig_desc = []
for df in [desc_cat, desc_cont]:
    sig_desc += [significant_values(df, 'indep_var', 'p_value')]

    #merge lists into one
sig_d = []
for i in sig_desc:
    sig_d += [*i]

    #merge all variables to use into one dataframe
variables = continuous_vars+binary_vars+categorical_vars

    #Get univariate odds ratios
uni_OR = []
for var in variables:
    uni_OR += [log_model(data_imp, 'treated', var, 'not_adjusted')]

uni_var = pd.concat(uni_OR) #create dataframe with all results
uni_var['variables'] = uni_var.index #rename index column

#Select variables to include in multivariate analysis

    # Identify significant univariate variables
sig_ls = significant_values(uni_var,'variables','p-value_not_adjusted')
    
    # Drop constant columns
sig_ls = [col for col in sig_ls if data_imp[col].nunique() > 1]

    #Get multivariate odds ratios
adj_var = log_model(data_imp, 'treated', sig_ls, 'adjusted')

    # Merge univariate and multivariate results
multivar = pd.merge(uni_var, adj_var, how='outer', left_index=True, right_index=True)

    # Add a 'significant' column based on adjusted p-values, if available
multivar['significant'] = np.where(
    multivar['p-value_adjusted'].notna() & (multivar['p-value_adjusted'] <= 0.05),
    'significant',
    'not significant'
)

# Export dataframe to excel
multivar.drop(columns=['variables']).to_excel(path+r'multivariate_analysis.xlsx')

"""Colour code results with significant results"""

wb = openpyxl.load_workbook('multivariate_analysis.xlsx')
name = wb.sheetnames
ws = wb[name[0]]

# Define red fill and bold red font for significant p values
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bold_red_font = Font(bold=True, color="9C0006")  # Dark red text
# ws = your worksheet object
significant_col_name = 'significant'  # column header with the "significant" string

header = [cell.value for cell in ws[1]]
if significant_col_name not in header:
    print(f"Column '{significant_col_name}' not found in header.")
else:
    col_idx = header.index(significant_col_name) + 1  # 1-based index

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        value = cell.value

        if isinstance(value, str) and value.strip().lower() == 'significant':
            for col in range(1, ws.max_column + 1):
                cell_to_format = ws.cell(row=row, column=col)
                cell_to_format.fill = red_fill
                cell_to_format.font = bold_red_font
                
# Save workbook with updated formatting
wb.save("multivariate_analysis.xlsx")
